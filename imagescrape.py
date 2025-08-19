import requests
from openpyxl import load_workbook
from pathlib import Path
from PIL import Image
import io

# === CONFIG ===
SHEET_ID = "1q01STuHSABJcptgHRmfFeXzkhFRtY6ymurO9uxkflfQ"
SHEET_GID = "0"  # not really needed since we export the whole workbook

# This URL downloads the sheet as an Excel file (whole workbook, all sheets)
EXCEL_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

# Output folders
output_folder1 = Path("thumbnails")
output_folder1.mkdir(exist_ok=True)

output_folder2 = Path("aill_thumbnails")
output_folder2.mkdir(exist_ok=True)

# === DOWNLOAD THE SHEET ===
print("Downloading sheet...")
response = requests.get(EXCEL_URL)
response.raise_for_status()

excel_file = "sheet.xlsx"
with open(excel_file, "wb") as f:
    f.write(response.content)

print("Downloaded sheet to", excel_file)

# === LOAD WORKBOOK ===
wb = load_workbook(excel_file)

# Function to extract images from a given worksheet
def extract_images(ws, output_folder):
    print(f"Processing sheet: {ws.title}")

    # Map row number to image
    image_map = {}
    for img in ws._images:  # internal API
        row_num = img.anchor._from.row + 1  # Excel rows are 1-indexed
        image_map[row_num] = img

    # Iterate over rows
    for row in range(2, 1000):  # D2:D999, A2:A999
        filename = ws[f"A{row}"].value
        if not filename:
            continue

        img = image_map.get(row)
        if not img:
            continue

        img_path = output_folder / f"{filename}.png"

        # Skip if file already exists
        if img_path.exists():
            print(f"Skipping {img_path} (already exists)")
            continue

        # Convert BytesIO to PIL Image
        image_data = img._data()  # returns BytesIO
        pil_img = Image.open(io.BytesIO(image_data))

        # Save image as PNG
        pil_img.save(img_path)
        print(f"Saved {img_path}")

# === Run extraction for first two sheets ===
if len(wb.worksheets) >= 1:
    extract_images(wb.worksheets[0], output_folder1)

if len(wb.worksheets) >= 2:
    extract_images(wb.worksheets[1], output_folder2)
